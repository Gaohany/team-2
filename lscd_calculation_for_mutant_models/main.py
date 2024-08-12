import argparse
from datetime import datetime
from metrics_utils.metrics_dataloader import metrics_dataloader
import toml
import importlib
from fuzzer.model_structure import model_structure
from metrics_utils.features_from_input import *
from metrics_utils.visualization_gini_index import *
from utils.util import obj
from pathlib import Path
import sys
import traceback
import openpyxl
COMPANY = "tum "
ALGORITHM = "gtsrb-new"
MODEL_VERSION = "v1"
TIME = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def get_all_result_files(directory):
        json_files = []
        for root, dirs, files in os.walk(directory):
            for file in files:
                json_files.append(root +"/"+ file)
        return json_files

def write_results_to_excel(file_name, model, directory_to_search):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[f"{model}"]
    json_files = get_all_result_files(directory_to_search)
    
    for i in range(0,len(json_files)):
        data_list = json_files[i].split("/")[-1].split("_")
        dataset = data_list[0]
        augmented_dataset = data_list[1]
        mutant_model = data_list[2]+ "_0"
        if mutant_model == "org.json_0":
            mutant_model = "AAA_Original"
        name_index = 1
        if augmented_dataset == "org":
            index = 1
            dataset_type = "org"
        elif augmented_dataset == "nc":
            index = 2
            dataset_type = "crashes"
        elif augmented_dataset ==  "kmnc":
            index = 3
            dataset_type = "crashes"
        elif augmented_dataset ==  "nbc":
            index = 4
            dataset_type = "crashes"
        elif augmented_dataset ==  "passed":
            index = 5
            dataset_type = "mt"
        elif augmented_dataset ==  "failed":
            index = 6
            dataset_type = "mt"
        else:
            print("index wrong")
            break
        try:
            for k in range(1,145, 6):
                if ws.cell(row= k, column = 2).value == mutant_model:
                    with open(json_files[i], "r") as f:
                        data = json.load(f)
                    ws[f'I{k+index-1}'] = data[f"LSCD_test_{dataset_type}"] 
                else:
                    continue
        except Exception as e:
            log_exceptions_to_file(e, f"exceptions_{model}_write_to_excel.log", json_files[i])
    wb.save(file_name)      


def load_module(module_name: str, filepath: Path):
    if module_name in sys.modules:
        return sys.modules[module_name]
    spec = importlib.util.spec_from_file_location(module_name, filepath.resolve()) 
    module = importlib.util.module_from_spec(spec) 
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    print("Loaded module", module_name)

    return module


def load_model(module, device, network_type):
    #Load Mutant Models from trained mutant models' folders for init centroids calculation
    model_name = module.__name__.split(".")[0]
    print(model_name)
    if network_type == "mnist":
        model_mt = module.MNIST_Lenet5()
    elif network_type == "gtsrb":
        model_mt = module.GTSRB_new()
    model_mt.model_name = model_name
    model_mt.model_path = Path(f'{args.mt_models_dir}/trained_mutants/{model_name}_000/model.pth')
    model_mt.load_state_dict(torch.load(model_mt.model_path, map_location=device))
    model_mt.to(device)
    model_mt.eval()
    
    return model_mt

def log_exceptions_to_file(exception, file_path, model, item):
    #save all exceptions for analysis
    with open(file_path, 'a') as f:
        traceback_str = traceback.format_exc()
        f.write(f"{item}  {model.model_name}")
        f.write("\n")
        f.write(traceback_str)
        f.write("\n\n\n") 
    return None


def arguments():
    parser = argparse.ArgumentParser(description="Dataset Quality Metrics")
    parser.add_argument("-output_dir", help="output directory to store results", default="results/gtsrb/")
    parser.add_argument(
        "-metrics_config_file",
        help="choose metrics configuration with which to run",
        default="config/dataset_metrics_config.toml",
    )
    parser.add_argument(
        "-dataset_config_file", help="choose dataset configuration with which to run", default="config/gtsrb.toml"
    )
    parser.add_argument(
        "-mt_models_dir", help="the directory which stores the mutant models", default = "./gtsrb"
    )
    return parser.parse_args()


if __name__ == "__main__": 

    dataset_criteria = ["org", "nc", "kmnc", "nbc", "passed","failed"]
    error_list = []

    args = arguments()
    metrics_config = json.loads(json.dumps(toml.load(args.metrics_config_file)), object_hook=obj)
    config = json.loads(json.dumps(toml.load(args.dataset_config_file)), object_hook=obj)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print("Using device:", device)
    config.device = device
    torch.set_grad_enabled(False)

    folders = sorted(
        [
            folder
            for folder in (Path(args.mt_models_dir) / "raw_mutants").iterdir()
            if folder.is_dir()
        ]
    )
    mutant_module = [load_module(".".join([folder.name, "model"]), folder / "model.py") for folder in folders]
    mutant_module.insert(0, "org")
    
    if not os.path.exists("saved_parameters"):
        os.makedirs(os.path.join("saved_parameters", "init_centroids"))

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    header = dict(
        __version_entry__=[
            dict(__Tool__="{}{}".format(COMPANY, ALGORITHM), __Version__="{}".format(MODEL_VERSION), __Time__=TIME)
        ]
    )

    for item in dataset_criteria:
        dataset_type = ""
        if item == "org":
            dataset_type = "org"
        elif item == "nc":
            dataset_type = "crashes"
            config.experiment_path = (
                ["E:/datasets/gtsrb_1_nc/gtsrb_1_nc/"] # Change to the path of corresponding dataset
            )
        elif item == "kmnc":
            dataset_type = "crashes"
            config.experiment_path = (
                ["E:/datasets/gtsrb_1_kmnc/gtsrb_1_kmnc/"] # Change to the path of corresponding dataset
            )
        elif item == "nbc":
            dataset_type = "crashes"
            config.experiment_path = (
                ["E:/datasets/gtsrb_1_nbc/gtsrb_1_nbc/"] # Change to the path of corresponding dataset
            )  
        elif item == "passed":
            dataset_type = "mt"
            config.experiment_path = (
                ["E:/datasets/gtsrb_mt/passed"] # Change to the path of corresponding dataset
            )  
        elif item == "failed":
            dataset_type = "mt"
            config.experiment_path = (
                ["E:/datasets/gtsrb_mt/failed"] # Change to the path of corresponding dataset
            ) 

        print("Storing feature vectors for the given {} dataset once.".format(dataset_type))

        dataset_test = metrics_dataloader(config, dataset_type,mode=metrics_config.mode)

        for module in mutant_module:
            if module == "org":
                if config.model == "mnist-lenet5":
                    model = model_structure[config.model]()
                elif config.model == "gtsrb-new":
                    model = model_structure[config.model]()
                elif config.model == "gtsrb-lenet":
                    model = model_structure[config.model]()
                elif config.model == "svhn-mixed":
                    model = model_structure[config.model]()
                else:
                    raise NotImplementedError("Model specified has not been implemented.")
                
                model.load_state_dict(torch.load(model.model_path, map_location=device))
                model.to(device)
                model.eval()
                log_file = os.path.join(args.output_dir, f"{config.detection_model.dataset}_{item}_{module}.json")

                if os.path.exists(log_file):
                    print("Using %s for logging informations." % (log_file))
                else:
                    with open(log_file, "w") as json_file:
                        json.dump(header, json_file, indent=4)
                        print("Results written to:", log_file)
            else:
                model = load_model(module, device, config.detection_model.image_set)
                log_file = os.path.join(args.output_dir, f"{config.detection_model.dataset}_{item}_{model.model_name}.json")
                if os.path.exists(log_file):
                    print("Using %s for logging informations." % (log_file))
                else:
                    with open(log_file, "w") as json_file:
                        json.dump(header, json_file, indent=4)
                        print("Results written to:", log_file)
            centroids_save_path = f"saved_parameters/init_centroids/{config.detection_model.dataset}/{model.model_name}.pickle"
            if model.model_name not in error_list:
                # Calculate initial centroid positioning using training dataset.
                try:
                    if not os.path.exists(centroids_save_path):
                        calculate_initial_centroid_radius(
                            model,
                            config,
                            centroids_save_path=centroids_save_path,
                            log_filepath=log_file,
                        )
                    else:
                        print("Initial centroids & radius values are used from:", centroids_save_path)

                    print("Calculating dataset metrics for {} {}.".format(item, model.model_name))

                    if item == "org":
                        vectors_dir = os.path.join("vectors", f"{config.data}", "_".join([item, model.model_name]))
                    else:
                        vectors_dir = os.path.join("vectors", f"{config.data}", "_".join([item, model.model_name]))

                    if not os.path.exists(vectors_dir):
                        os.makedirs(vectors_dir)
                    feature_vector_file_name = os.path.join(vectors_dir, f"vectors.pickle")

                    if os.path.exists(feature_vector_file_name):
                        with open(feature_vector_file_name, "rb") as f:
                            print("Loading feature vectors from: ", feature_vector_file_name)
                            feature_dict = torch.load(f)
                    else:
                        # Create and store feature vectors for given dataset and type of images.
                        store_feature_vectors_mt(
                            model,
                            config,
                            dataset_type,
                            mode=metrics_config.mode,
                            output_file_name=feature_vector_file_name,
                            dataset_test = dataset_test,
                        )

                        if os.path.exists(feature_vector_file_name):
                            with open(feature_vector_file_name, "rb") as f:
                                print("Loading feature vectors from: ", feature_vector_file_name)
                                feature_dict = torch.load(f)
                    #Calculate LSCD
                    lscd_avg, avg_classes_distance, max_classes_distance = calculate_lscd(
                        model,
                        config,
                        centroids_save_path=centroids_save_path,
                        log_filepath=log_file,
                        feature_vectors_dict_main=feature_dict,
                        type=dataset_type,
                        radius_type=metrics_config.lscd_radius_type,
                        output_path=feature_vector_file_name,
                    )
                    print("Latent Space Class Dispersion Values (in %): \n", lscd_avg)
                    print("Average Latent Space Class Distances: \n", avg_classes_distance)
                    print("Maximum Latent Space Class Distances: \n", max_classes_distance)
                    
                    del model

                except Exception as e:
                    if model.model_name not in error_list:
                        error_list.append(model.model_name)
                    log_exceptions_to_file(e, f"exceptions_{config.detection_model.dataset}.log", model, item)
            else:
                continue
    