from pathlib import Path
from scipy.stats import spearmanr, kendalltau
from sklearn.metrics import r2_score
from ms_calculations.mutagen.ms_utils import *
import pandas as pd
import openpyxl
import argparse
import toml
from all_classes import *

#sys.path.append("E:/Code/fdr_ms_calculations_liu/fdr_ms_calculations/mutagen")

parser = argparse.ArgumentParser()
parser.add_argument(
    "-result_dir",
    default=Path("./results/mnist"),
    help="Result Path",
)
parser.add_argument(
    "-config_file",
    help="choose configuration with which to run",
    default="../config/mnist.toml",
)



config = json.loads(json.dumps(toml.load(parser.config_file)), object_hook=obj)

model = config.data
results_directory = Path(parser.result_dir)

selected_splits = ["test", "cc_nc", "cc_kmnc", "cc_nbc", "passed", "failed"]
selected_splits_analyze = ["test", "cc_nc", "cc_kmnc", "cc_nbc","passed", "failed"]
num_classes = config.num_classes

trained_mutant_folders = sorted(
    [
        folder
        for folder in (results_directory / "trained_mutants").iterdir()
        if folder.is_dir()
    ]
)

operator_list = [
    str(selected_mutant).split("/")[-1] for selected_mutant in trained_mutant_folders
]

for split in selected_splits:
    evaluation_folders = sorted(
    [
        folder
        for folder in (results_directory / f"evaln/dataset={split}").iterdir()
        if folder.is_dir()
    ]
    )

    selected_folder = evaluation_folders[0]


    parquet_file_path = Path(
        selected_folder, "sut_training=0", f"{split}.parquet"
    )

    org_data = pd.read_parquet(parquet_file_path, engine="fastparquet")

    acc_ref = round(
            (org_data["label"] == org_data["output"]).sum()
            / org_data.shape[0]
            * 100,
            2,
        )

evaluation_folders_updated = {}
for split in selected_splits:
    evaluation_folders = sorted(
    [
        folder
        for folder in (results_directory / f"evaln/dataset={split}").iterdir()
        if folder.is_dir()
    ]
    )    
    evaluation_folders_updated.update({split: evaluation_folders})

# To calculate MS for all mutatnts in all splits...
ms_dict, lscd_dict, acc_dict, pass_rate_dict = {}, {}, {}, {}
index_last, index_current = 0, 0

for split in selected_splits_analyze:
    ms_list, lscd_list, scaled_lscd_list, acc_list, pass_rate_list= [], [], [], [], []

    evaluation_folders_current = evaluation_folders_updated[split]

    for i, selected_folder in enumerate(evaluation_folders_current):

        parquet_file_path = Path(
            selected_folder, "sut_training=0", f"{split}.parquet"
        )
        ref_parquet_file_path = Path(
            evaluation_folders_current[0], "sut_training=0", f"{split}.parquet"
        ) 
        input_data = pd.read_parquet(parquet_file_path, engine="fastparquet")
        ref_data = pd.read_parquet(ref_parquet_file_path, engine="fastparquet")

        mutation_score, cl_ms_score_dict = calculate_mutation_score(
            input_data=input_data, reference_data=ref_data, num_classes=num_classes, split=split
        )

        accuracy = round(
            (input_data["label"] == input_data["output"]).sum()
            / input_data.shape[0]
            * 100,
            2,
        )
        if split == "test":
            pass_rate = round(
                (input_data["output"] == input_data["output"]).sum()
                / input_data.shape[0]
                * 100,
                2,
            )
        else:
            pass_rate = round(
                (input_data["output"] == input_data[f"ori_output"]).sum()
                / input_data.shape[0]
                * 100,
                2,
            )
        acc_list.append(accuracy)
        pass_rate_list.append(pass_rate)
        ms_list.append([input_data["sut_name"][0], round(mutation_score,3)])
    ms_dict.update({split: ms_list})

    acc_dict.update({split: acc_list})
    pass_rate_dict.update({split: pass_rate_list})

file_name = f"results.xlsx"
wb = openpyxl.load_workbook(file_name)
ws = wb[f"{model}"]

for dataset in selected_splits_analyze:
    ms_result = ms_dict[dataset]
    acc_result = acc_dict[dataset]
    pass_rate = pass_rate_dict[dataset]
    name_index = 1
    print(dataset)
    if dataset == "test":
        index = 1
    elif dataset == "cc_nc":
        index = 2
    elif dataset ==  "cc_kmnc":
        index = 3
    elif dataset ==  "cc_nbc":
        index = 4
    elif dataset ==  "passed":
        index = 5
    elif dataset ==  "failed":
        index = 6
    else:
        print("index wrong")
        break
    for i in range(0,len(ms_result)):

        ws[f'B{name_index}'] = ms_result[i][0]
        ws[f'E{index}'] = acc_result[i]
        ws[f'F{index}'] = pass_rate[i]
        ws[f'G{index}'] = ms_result[i][1]
        name_index = name_index + 6
        index = index + 6
wb.save(file_name)

